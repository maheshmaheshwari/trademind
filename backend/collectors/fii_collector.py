"""
TradeMind AI — FII/DII Data Collector

Collects FII (Foreign Institutional Investor) and DII (Domestic Institutional
Investor) net buy/sell data and stores it day-wise in `fii_dii_daily`.

Modes
─────
1. Daily live       — collect_fii_dii_data()     → today from NSE live API
2. Historical CSV   — import_from_csv(path)       → bulk import from NSE Excel/CSV
3. Backfill script  — backfill_fii_dii(years=3)  → tries multiple free sources

Historical data sources (in priority order)
────────────────────────────────────────────
A. NSE India manual download:
   https://www.nseindia.com/reports/fii-dii-activity
   → Download "FII / DII Statistics" Excel → pass path to import_from_csv()

B. Stooq — free Polish financial data provider with NSE FII proxy data
   https://stooq.com/q/d/l/?s=fiidii.in&i=d

C. NSE live API (today only — date param ignored by NSE's API)
   https://www.nseindia.com/api/fiidiiTradeReact
"""

import csv
import io
import logging
import time
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional

import requests

from database.db import get_connection, release_connection, _execute

logger = logging.getLogger(__name__)

_NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.nseindia.com/",
}
_NSE_FII_URL = "https://www.nseindia.com/api/fiidiiTradeReact"
_NSE_HOME    = "https://www.nseindia.com/"


# ── DB helpers ─────────────────────────────────────────────────────────────────

def _upsert_rows(rows: List[Dict]) -> None:
    if not rows:
        return
    conn = get_connection()
    try:
        for r in rows:
            _execute(conn, """
                INSERT INTO fii_dii_daily
                    (date, fii_net, dii_net, fii_buy, fii_sell,
                     dii_buy, dii_sell, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (date) DO UPDATE SET
                    fii_net  = EXCLUDED.fii_net,
                    dii_net  = EXCLUDED.dii_net,
                    fii_buy  = EXCLUDED.fii_buy,
                    fii_sell = EXCLUDED.fii_sell,
                    dii_buy  = EXCLUDED.dii_buy,
                    dii_sell = EXCLUDED.dii_sell,
                    source   = EXCLUDED.source
            """, (
                r["date"],
                r.get("fii_net", 0),  r.get("dii_net", 0),
                r.get("fii_buy",  0), r.get("fii_sell", 0),
                r.get("dii_buy",  0), r.get("dii_sell", 0),
                r.get("source", "nse"),
            ))
        conn.commit()
        logger.info("Upserted %d FII/DII rows", len(rows))
    finally:
        release_connection(conn)


def _existing_dates() -> set:
    conn = get_connection()
    try:
        cur = _execute(conn, "SELECT date FROM fii_dii_daily", ())
        return {str(r[0]) for r in cur.fetchall()}
    finally:
        release_connection(conn)


# ── NSE session ────────────────────────────────────────────────────────────────

def _nse_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(_NSE_HEADERS)
    try:
        s.get(_NSE_HOME, timeout=12)
    except Exception:
        pass
    return s


def _parse_nse_response(data: list) -> Dict:
    fii_buy = fii_sell = dii_buy = dii_sell = 0.0
    for entry in data:
        cat  = str(entry.get("category", "")).upper()
        buy  = float(entry.get("buyValue",  0) or 0)
        sell = float(entry.get("sellValue", 0) or 0)
        if "FII" in cat or "FPI" in cat:
            fii_buy += buy; fii_sell += sell
        elif "DII" in cat:
            dii_buy += buy; dii_sell += sell
    return {
        "fii_net":  round(fii_buy  - fii_sell, 2),
        "dii_net":  round(dii_buy  - dii_sell, 2),
        "fii_buy":  round(fii_buy,  2),
        "fii_sell": round(fii_sell, 2),
        "dii_buy":  round(dii_buy,  2),
        "dii_sell": round(dii_sell, 2),
    }


# ── Mode 1: Daily live collection ─────────────────────────────────────────────

def collect_fii_dii_data() -> Optional[Dict]:
    """
    Fetch TODAY's FII/DII data from NSE live API and upsert into fii_dii_daily.
    Called by the scheduler every weekday at 17:00 IST.
    """
    try:
        session = _nse_session()
        resp = session.get(_NSE_FII_URL, timeout=12)
        if resp.status_code != 200:
            logger.warning("NSE FII/DII API: %s", resp.status_code)
            return None

        data = resp.json()
        # NSE returns latest available date — use that date, not today's
        if data:
            raw_date = data[0].get("date", "")
            try:
                d = datetime.strptime(raw_date, "%d-%b-%Y").strftime("%Y-%m-%d")
            except ValueError:
                d = datetime.now().strftime("%Y-%m-%d")
        else:
            d = datetime.now().strftime("%Y-%m-%d")

        parsed = _parse_nse_response(data)
        parsed["date"]   = d
        parsed["source"] = "nse_live"
        _upsert_rows([parsed])
        logger.info("FII/DII today (%s): FII=%.2f Cr  DII=%.2f Cr",
                    d, parsed["fii_net"], parsed["dii_net"])
        return parsed

    except Exception as exc:
        logger.error("collect_fii_dii_data failed: %s", exc)
        return None


# ── Mode 2: CSV / Excel import (manual one-time historical load) ───────────────

def import_from_csv(filepath: str) -> int:
    """
    Import historical FII/DII data from a CSV/Excel file.

    NSE provides historical FII/DII data as downloadable Excel files from:
      https://www.nseindia.com/reports/fii-dii-activity

    Expected CSV columns (flexible — tries multiple header names):
      Date, FII Buy, FII Sell, FII Net, DII Buy, DII Sell, DII Net

    Returns number of rows inserted.
    """
    import os

    rows: List[Dict] = []

    # Try reading as Excel first, then CSV
    if filepath.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = dict(zip(headers, row))
                parsed = _parse_row_dict(r)
                if parsed:
                    rows.append(parsed)
        except ImportError:
            logger.error("openpyxl not installed — install with: pip install openpyxl")
            return 0
    else:
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r in reader:
                parsed = _parse_row_dict({k.strip().lower(): v for k, v in r.items()})
                if parsed:
                    rows.append(parsed)

    if not rows:
        logger.warning("No valid rows found in %s", filepath)
        return 0

    existing = _existing_dates()
    new_rows = [r for r in rows if r["date"] not in existing]
    _upsert_rows(new_rows)
    logger.info("CSV import: %d rows total, %d new", len(rows), len(new_rows))
    return len(new_rows)


def _parse_row_dict(r: dict) -> Optional[Dict]:
    """Parse a dict of CSV/Excel column values into a FII/DII row."""
    # Try various date column names
    date_val = (r.get("date") or r.get("trade date") or
                r.get("tradedate") or r.get("mdate") or "")
    if not date_val:
        return None

    date_str = _normalise_date(str(date_val).strip())
    if not date_str:
        return None

    def _f(keys):
        for k in keys:
            v = r.get(k)
            if v not in (None, "", "-", "N/A"):
                try:
                    return float(str(v).replace(",", ""))
                except ValueError:
                    pass
        return 0.0

    fii_buy  = _f(["fii buy", "fii_buy", "fii purchase", "fii gross buy"])
    fii_sell = _f(["fii sell", "fii_sell", "fii sale", "fii gross sell"])
    fii_net  = _f(["fii net", "fii_net", "fii net purchase / sale"])
    dii_buy  = _f(["dii buy", "dii_buy", "dii purchase", "dii gross buy"])
    dii_sell = _f(["dii sell", "dii_sell", "dii sale", "dii gross sell"])
    dii_net  = _f(["dii net", "dii_net", "dii net purchase / sale"])

    # If buy/sell not available but net is, keep net; otherwise compute net
    if fii_buy == 0 and fii_sell == 0 and fii_net != 0:
        pass
    else:
        fii_net = fii_net or round(fii_buy - fii_sell, 2)

    if dii_buy == 0 and dii_sell == 0 and dii_net != 0:
        pass
    else:
        dii_net = dii_net or round(dii_buy - dii_sell, 2)

    return {
        "date":     date_str,
        "fii_net":  round(fii_net, 2),
        "dii_net":  round(dii_net, 2),
        "fii_buy":  round(fii_buy, 2),
        "fii_sell": round(fii_sell, 2),
        "dii_buy":  round(dii_buy, 2),
        "dii_sell": round(dii_sell, 2),
        "source":   "csv_import",
    }


def _normalise_date(s: str) -> Optional[str]:
    """Try to parse a date string into YYYY-MM-DD."""
    formats = ["%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y",
               "%b %d, %Y", "%d %b %Y", "%d-%b-%y", "%d/%m/%y"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ── Mode 3: Automated backfill ────────────────────────────────────────────────

def backfill_fii_dii(years: int = 3) -> int:
    """
    Backfill fii_dii_daily using available free sources.

    Priority:
    1. Stooq.com — provides Indian market FII/DII proxy data as CSV
    2. Today's NSE live data (adds just today if everything else fails)

    For a complete 3-year history, the most reliable method is manual:
      1. Visit https://www.nseindia.com/reports/fii-dii-activity
      2. Download the FII/DII Statistics Excel file
      3. Run: python -c "from collectors.fii_collector import import_from_csv; import_from_csv('path/to/file.xlsx')"

    Returns number of rows inserted.
    """
    logger.info("Starting FII/DII backfill (target: %d years)…", years)
    cutoff   = date.today() - timedelta(days=years * 365)
    existing = _existing_dates()
    inserted = 0

    # ── Try Stooq ─────────────────────────────────────────────────────────────
    stooq_rows = _fetch_stooq(cutoff)
    if stooq_rows:
        new_rows = [r for r in stooq_rows if r["date"] not in existing]
        if new_rows:
            _upsert_rows(new_rows)
            inserted += len(new_rows)
            existing.update(r["date"] for r in new_rows)
            logger.info("  Stooq: inserted %d rows", len(new_rows))
        else:
            logger.info("  Stooq: already up to date")
    else:
        logger.info("  Stooq: no data returned (might not have Indian FII/DII data)")

    # ── Add today from NSE live ────────────────────────────────────────────────
    today_data = collect_fii_dii_data()
    if today_data and today_data["date"] not in existing:
        inserted += 1

    if inserted == 0:
        logger.warning(
            "Automated backfill returned 0 rows.\n"
            "For a full 3-year history, manually download NSE FII/DII data:\n"
            "  1. Go to https://www.nseindia.com/reports/fii-dii-activity\n"
            "  2. Download 'FII / DII Statistics' as Excel/CSV\n"
            "  3. Run: python -c \"from collectors.fii_collector import import_from_csv; "
            "import_from_csv('path/to/file.xlsx')\""
        )

    logger.info("Backfill complete — %d rows upserted", inserted)
    return inserted


def _fetch_stooq(cutoff: date) -> List[Dict]:
    """
    Fetch FII/DII data from Stooq (free financial data CDN).
    Stooq provides Indian market data including institutional flow proxies.
    """
    try:
        from_str = cutoff.strftime("%Y%m%d")
        to_str   = date.today().strftime("%Y%m%d")
        # Stooq CSV for FII net buy/sell proxy
        url = f"https://stooq.com/q/d/l/?s=fiidii.in&d1={from_str}&d2={to_str}&i=d"
        resp = requests.get(url, timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code != 200 or "No data" in resp.text[:50]:
            logger.debug("Stooq: %s / empty", resp.status_code)
            return []

        rows = []
        reader = csv.DictReader(io.StringIO(resp.text))
        for rec in reader:
            date_str = rec.get("Date", "").strip()
            if not date_str:
                continue
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                continue
            if d < cutoff:
                continue

            close = float(rec.get("Close", 0) or 0)
            open_ = float(rec.get("Open",  0) or 0)
            # Stooq FII/DII index: positive = net buy
            rows.append({
                "date":     str(d),
                "fii_net":  round(close, 2),      # close = net FII flow proxy
                "dii_net":  0.0,                  # DII not separately available
                "fii_buy":  max(close, 0),
                "fii_sell": max(-close, 0),
                "dii_buy":  0.0,
                "dii_sell": 0.0,
                "source":   "stooq",
            })
        logger.info("Stooq returned %d FII/DII records", len(rows))
        return rows

    except Exception as exc:
        logger.debug("Stooq fetch failed: %s", exc)
        return []


# ── DB stats helper ────────────────────────────────────────────────────────────

def get_fii_dii_stats() -> Dict:
    """Return summary stats for the fii_dii_daily table."""
    conn = get_connection()
    try:
        cur = _execute(conn, """
            SELECT COUNT(*), MIN(date), MAX(date),
                   AVG(fii_net), AVG(dii_net)
            FROM fii_dii_daily
        """, ())
        row = cur.fetchone()
        return {
            "rows":     row[0] or 0,
            "min_date": str(row[1]) if row[1] else None,
            "max_date": str(row[2]) if row[2] else None,
            "avg_fii":  round(float(row[3] or 0), 2),
            "avg_dii":  round(float(row[4] or 0), 2),
        }
    finally:
        release_connection(conn)


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")

    cmd = sys.argv[1] if len(sys.argv) > 1 else "today"

    if cmd == "backfill":
        years = int(sys.argv[2]) if len(sys.argv) > 2 else 3
        n = backfill_fii_dii(years=years)
        print(f"Backfill: {n} rows inserted/updated")

    elif cmd == "import":
        if len(sys.argv) < 3:
            print("Usage: python fii_collector.py import <path/to/file.csv>")
        else:
            n = import_from_csv(sys.argv[2])
            print(f"Imported {n} rows")

    elif cmd == "stats":
        stats = get_fii_dii_stats()
        print(f"fii_dii_daily table stats:")
        for k, v in stats.items():
            print(f"  {k}: {v}")

    else:
        data = collect_fii_dii_data()
        if data:
            print(f"FII Net: ₹{data['fii_net']} Cr  |  DII Net: ₹{data['dii_net']} Cr  |  Date: {data['date']}")
        else:
            print("Could not fetch today's FII/DII data")
